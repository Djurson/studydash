import openpyxl
import json
import argparse

# Argument parsing
parser = argparse.ArgumentParser()
parser.add_argument("-i", "--input", default="data_processed.xlsx",
                    help="Input Excel file", metavar="file")
parser.add_argument("-o", "--output", default="data_processed.json",
                    help="Output JSON file", metavar="file")
options = parser.parse_args()

# Load the Excel file
wb = openpyxl.load_workbook(options.input)
sheet = wb.active

# Extract header row
header = next(sheet.iter_rows(values_only=True))

# Extract data rows and store them as dictionaries
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    entry = {header[i]: row[i] for i in range(len(header)) if header[i] is not None}
    data.append(entry)

# Organize data into hierarchical structure
course_structure = {"Masters": {}}
for entry in data:
    course_name = entry.pop("Namn", "Unknown Course")
    termin = entry.pop("Termin", None)
    period = entry.pop("Period", None)
    
    if termin is None:
        course_structure["Masters"][course_name] = entry
    else:
        termin_key = f"Termin {int(termin)}"
        if termin_key not in course_structure:
            course_structure[termin_key] = {}
        
        # Handle multiple periods or missing values
        if period is None:
            period_key = "No Period"
        elif isinstance(period, (int, float)):
            period_key = f"Period {int(period)}"
        else:
            period_key = f"Period {period}"  # Keep as string if not a number
        
        if period_key not in course_structure[termin_key]:
            course_structure[termin_key][period_key] = {}
        
        course_structure[termin_key][period_key][course_name] = entry

# Save structured data to a JSON file
with open(options.output, 'w', encoding='utf-8') as json_file:
    json.dump(course_structure, json_file, ensure_ascii=False, indent=4)

print(f"Data successfully converted to hierarchical JSON and saved to {options.output}")
