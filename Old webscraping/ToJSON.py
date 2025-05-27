import openpyxl
import json
import argparse

# Argument parsing
parser = argparse.ArgumentParser()
parser.add_argument("-i", "--input", default="data_processed.xlsx",
                    help="Input Excel file", metavar="file")
parser.add_argument("-o", "--output", default="data_processed.json",
                    help="Output JSON file", metavar="file")
options = parser.parse_args()

# Load the Excel file
wb = openpyxl.load_workbook(options.input)
sheet = wb.active

# Extract header row
header = next(sheet.iter_rows(values_only=True))

# Extract data rows and store them as dictionaries
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    entry = {header[i]: row[i] for i in range(len(header)) if header[i] is not None}
    data.append(entry)

# Save data to a JSON file
with open(options.output, 'w', encoding='utf-8') as json_file:
    json.dump(data, json_file, ensure_ascii=False, indent=4)

print(f"Data successfully converted to JSON and saved to {options.output}")
