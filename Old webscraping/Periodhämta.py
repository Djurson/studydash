import argparse
from openpyxl import load_workbook, Workbook

def parse_arguments():
    parser = argparse.ArgumentParser(description="Process Excel course data")
    parser.add_argument("-i", "--input", default="P6CMEN.xlsx", help="Input XLSX file")
    parser.add_argument("-o", "--output", default="output.xlsx", help="Output XLSX file")
    parser.add_argument("-p", "--program", help="Filter by program")
    parser.add_argument("-t", "--termin", help="Filter by term range, format N-M")
    parser.add_argument("--område", help="Filter by subject area")
    parser.add_argument("--one-course-per-cell", action="store_true", help="One course per cell")
    parser.add_argument("--show-code", action="store_true", help="Include course code")
    parser.add_argument("--show-name", action="store_true", help="Include course name")
    parser.add_argument("--show-campus", action="store_true", help="Include campus")
    parser.add_argument("--show-field", action="store_true", help="Include subject area")
    parser.add_argument("-c", "--compress", action="store_true", help="Compress term view")
    parser.add_argument("-v", "--vof", action="store_true", help="Include VOF info")
    return parser.parse_args()

def load_data(sheet, options):
    course_data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        (_, _, link_text, link_url, summary, program, term, period,
         block, _, campus, vof, _, name, _, _, _, _, _, _, _, _, syllabus) = row[:23]

        if not term or term == "null":
            continue
        if options.program and options.program != program:
            continue

        term_num = int(term.split()[0])
        if options.termin:
            min_term, max_term = map(int, options.termin.split("-"))
            if not (min_term <= term_num <= max_term):
                continue

        code = link_url.split("/")[4] if link_url else "UNKNOWN"

        if code not in course_data:
            course_data[code] = {
                "name": link_text,
                "subject": extract_field(summary, syllabus, "Huvudområde"),
                "sessions": set()
            }

        blocks = block.split(",")
        periods = period.split(",")
        for b, p in zip(blocks, periods):
            course_data[code]["sessions"].add((term_num, int(p), b.strip(), vof, campus))

    return course_data

def extract_field(summary, syllabus, key):
    # Placeholder for parsing logic from summary/syllabus
    # Just return "NA" for now, unless specified.
    return "NA"

def write_output(course_data, options, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed"

    headers = ["Termin", "Period", "-", "1", "2", "3", "4"]
    ws.append(headers)

    for row_idx in range(30):
        term = row_idx // 2
        period = row_idx % 2 + 1
        term_label = "HT" if options.compress and term else ("VT" if options.compress else term)

        row_data = {"-": [], "1": [], "2": [], "3": [], "4": []}
        data_written = False

        for code, data in course_data.items():
            if options.område and options.område not in data["subject"].split(", "):
                continue

            for session in data["sessions"]:
                sess_term, sess_period, block, vof, campus = session

                if options.compress:
                    if sess_term % 2 != term:
                        continue
                else:
                    if sess_term != term:
                        continue

                if sess_period != period:
                    continue

                text = ""
                if options.vof:
                    text += vof + " "
                if options.show_code:
                    text += code + " "
                if options.show_name:
                    text += data["name"] + " "
                if options.show_field:
                    text += f"({data['subject']}) "
                if options.show_campus:
                    text += f"({campus[0]})" if campus else ""

                for sub_block in block.split("+"):
                    row_data[sub_block].append(text.strip())

                data_written = True
                break

        if not data_written:
            continue

        if options.one_course_per_cell:
            for i in range(max(len(c) for c in row_data.values())):
                row = [term_label, period]
                for k in ["-", "1", "2", "3", "4"]:
                    row.append(row_data[k][i] if i < len(row_data[k]) else "")
                ws.append(row)
        else:
            row = [term_label, period]
            for k in ["-", "1", "2", "3", "4"]:
                row.append("\n".join(row_data[k]))
            ws.append(row)

    wb.save(output_file)

def main():
    args = parse_arguments()
    wb = load_workbook(args.input)
    sheet = wb.active
    course_data = load_data(sheet, args)
    write_output(course_data, args, args.output)

if __name__ == "__main__":
    main()
