#!/usr/bin/env python3

import argparse
import openpyxl
import csv
from openpyxl import Workbook
import hämta

parser = argparse.ArgumentParser()
parser.add_argument("-i", "--input", default="p6CMEN.xlsx",
                    help="Raw data file, from web scraping studieinfo.liu.se, in XLSX or CSV format", metavar="file")
parser.add_argument("-o", "--output", default="data_processed.xlsx",
                    help="Output file in XLSX or CSV format", metavar="file")
parser.add_argument("-p", "--program",
                    help="Visa bara för detta program", metavar="program")
parser.add_argument("-t", "--termin",
                    help="Visa bara dessa terminer", metavar="N-M")
parser.add_argument("--område",
                    help="Visa bara detta huvudområde", metavar="område")
parser.add_argument("--one-course-per-cell",
                    help="Skapa flera rader per period så att varje cell bara innehåller en kurs",
                    action='store_true', default=False)
parser.add_argument("--show-code",
                    help="Visa kurskoden",
                    action='store_true', default=False)
parser.add_argument("--show-name",
                    help="Visa kursnamnet",
                    action='store_true', default=False)
parser.add_argument("--show-campus",
                    help="Visa på vilket campus kursen ges",
                    action='store_true', default=False)
parser.add_argument("--show-field",
                    help="Visa inom vilket eller vilka huvudområden kursen ges",
                    action='store_true', default=False)
parser.add_argument("-c", "--compress",
                    help="Komprimera till HT/VT",
                    action='store_true', default=False)
parser.add_argument("-v", "--vof",
                    help="Visa VOF",
                    action='store_true', default=False)
options = parser.parse_args()

if options.termin is not None:
    if "-" not in options.termin:
        print("Felaktigt spann för termin")
        exit(1)
    span = options.termin.split("-")
    if len(span) != 2:
        print("Felaktigt spann för termin")
        exit(1)
    options.termin = (int(span[0]), int(span[1]))

if options.input.endswith(".xlsx"):
    # Handling XLSX input
    courses = {}

    wb = openpyxl.load_workbook(options.input)
    sheet = wb.active

    first = True
    for row in sheet.iter_rows(values_only=True):
        if first:
            first = False
            continue

        web_scraper_order, web_scraper_start_url, course_link, course_link_href, oversikt_grunddata, Kursen_ges_för, Termin, Period, Block, Språk, Ort, VOF, EXA, Benämning, Omfattning, Betygsskala, Name, Introducera, Undervisa, Anvanda, Moduler, Kommentar, kursplan = row

        kurskod = course_link_href.split("/")[4]

        if kurskod not in courses:
            courses[kurskod] = {"Namn": course_link}
            courses[kurskod].update(hämta.avsnitt(oversikt_grunddata))
            courses[kurskod].update(hämta.avsnitt(kursplan))

    output_wb = Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Processed Data"

    header = ["Kurs", "Namn", "Huvudområde", "Utbildningsnivå", "Kurstyp", "Examinator", "Studierektor", "Undervisningstid",
              "Fördjupningsnivå", "Rekommenderade förkunskaper", "Lärandemål", "Kursinnehåll", "Undervisnings- och arbetsformer",
              "Examination", "Betygsskala", "Övrig information", "Påbyggnadskurser", "Institution", "Kurslitteratur"]
    output_sheet.append(header)

    for kod, data in courses.items():
        row = [kod, data["Namn"], data.get("Huvudområde", ""), data.get("Utbildningsnivå", ""),
               data.get("Kurstyp", ""), data.get("Examinator", ""), data.get("Studierektor eller motsvarande", ""),
               data.get("Undervisningstid", ""), data.get("Fördjupningsnivå", ""), data.get("Rekommenderade förkunskaper", ""),
               data.get("Lärandemål", ""), data.get("Kursinnehåll", ""), data.get("Undervisnings- och arbetsformer", ""),
               data.get("Examination", ""), data.get("Betygsskala", ""), data.get("Övrig information", ""),
               data.get("Påbyggnadskurser", ""), data.get("Institution", ""), data.get("Kurslitteratur", "")]
        output_sheet.append(row)

    output_wb.save(options.output)

elif options.input.endswith(".csv"):
    # Handling CSV input
    courses = {}

    with open(options.input) as in_file:
        csv_reader = csv.reader(in_file)
        first = True
        for row in csv_reader:
            if first:
                first = False
                continue

            web_scraper_order, web_scraper_start_url, course_link, course_link_href, oversikt_grunddata, Kursen_ges_för, Termin, Period, Block, Språk, Ort, VOF, EXA, Benämning, Omfattning, Betygsskala, Name, Introducera, Undervisa, Anvanda, Moduler, Kommentar, kursplan = row

            if Termin == "" or Termin == "null":
                continue

            if options.program is not None and options.program != Kursen_ges_för:
                continue

            kurskod = course_link_href.split("/")[4]
            Termin = int(Termin.split()[0])

            if kurskod not in courses:
                courses[kurskod] = {"Namn": course_link, "Tillfälle": set()}
                courses[kurskod].update(hämta.avsnitt(oversikt_grunddata))
                courses[kurskod].update(hämta.avsnitt(kursplan))

            if not "," in Period:
                courses[kurskod]["Tillfälle"].add((Termin, int(Period), Block, False, VOF, Ort))
            else:
                for period_block in zip(Period.split(","), Block.split(",")):
                    courses[kurskod]["Tillfälle"].add((Termin, int(period_block[0]), period_block[1].strip(), True, VOF, Ort))

    with open(options.output, 'w') as out_file:
        csv_writer = csv.writer(out_file, quoting=csv.QUOTE_ALL)
        csv_writer.writerow(["Termin", "Period", "-", "1", "2", "3", "4"])

        for row_idx in range(0, 30):
            termin = row_idx // 2
            period = row_idx % 2 + 1

            row_courses = {'-': [], '1': [], '2': [], '3': [], '4': []}
            write_data = False

            for kod, data in courses.items():
                huvudområde = data["Huvudområde"].split(", ") if "Huvudområde" in data else []
                if options.område is not None and options.område not in huvudområde:
                    continue

                for tillf in data["Tillfälle"]:
                    if not options.termin is None and \
                            (tillf[0] < options.termin[0] or options.termin[1] < tillf[0]):
                        continue

                    if not options.compress:
                        if tillf[0] != termin:
                            continue
                    else:
                        if tillf[0] % 2 != termin:
                            continue

                    if tillf[1] != period:
                        continue

                    write_data = True

                    for block in tillf[2].split("+"):
                        txt = ""
                        if options.vof:
                            txt += tillf[4] + " "
                        if options.show_code:
                            txt += f"{kod}" + ("*" if tillf[3] else " ")
                        if options.show_name:
                            txt += f"""{data["Namn"]}"""
                        if options.show_campus:
                            txt += (f""" ({tillf[5][0]})""" if len(tillf[5]) > 0 else "")
                        if options.show_field:
                            txt += f""" ({data["Huvudområde"]})""" if "Huvudområde" in data else ""
                        row_courses[block].append(txt)

                    break

            if not options.compress:
                row = [termin, period]
            else:
                if termin:
                    row = ["HT", period]
                else:
                    row = ["VT", period]

            if options.one_course_per_cell:
                for idx in range(10):
                    found = False
                    for block, data in row_courses.items():
                        if len(data) > idx:
                            found = True
                            row.append(data[idx].strip())
                        else:
                            row.append("")
                    if found:
                        csv_writer.writerow(row)
                        row = ["", ""]
                        write_data = False
            else:
                for block, data in row_courses.items():
                    text = ""
                    for item in data:
                        text = text + item + "\n"
                    row.append(text.strip())

            if write_data:
                csv_writer.writerow(row)
